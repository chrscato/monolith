# billing/logic/postprocess/utils/excel_generator.py

import pandas as pd
import holidays
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import logging
import re
from decimal import Decimal

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Get the absolute path to the monolith root directory
DB_ROOT = Path(r"C:\Users\ChristopherCato\OneDrive - clarity-dx.com\code\monolith")

class ExcelBatchGenerator:
    """
    Generator for Excel batch files and historical log updates.
    Handles duplicate detection, EOBR numbering, and business day calculations.
    """
    
    def __init__(self, historical_excel_path: Path = None):
        """
        Initialize the Excel batch generator.
        
        Args:
            historical_excel_path: Path to the Historical_EOBR_Data.xlsx file
        """
        if historical_excel_path is None:
            historical_excel_path = Path(r"C:\Users\ChristopherCato\OneDrive - clarity-dx.com\code\monolith\billing\logic\postprocess\batch_outputs\Historical_EOBR_Data copy.xlsx")
        
        self.historical_excel_path = historical_excel_path
        self.historical_df = None
        self.us_holidays = None
        self.current_batch_keys = set()  # Track keys from current batch
        self._load_historical_data()
        self._init_holidays()
    
    def _load_historical_data(self):
        """Load historical Excel data or create empty DataFrame."""
        try:
            if self.historical_excel_path.exists():
                # Read Excel with explicit dtype handling for numeric columns
                self.historical_df = pd.read_excel(
                    self.historical_excel_path,
                    dtype={
                        'Release Payment': 'str',
                        'Duplicate Check': 'str', 
                        'Full Duplicate Key': 'str',
                        'Input File': 'str',
                        'EOBR Number': 'str',
                        'Vendor': 'str',
                        'Mailing Address': 'str',
                        'Terms': 'str',
                        'Bill Date': 'str',
                        'Due Date': 'str',
                        'Category': 'str',
                        'Description': 'str',
                        'Memo': 'str'
                        # Amount and Total will be converted below
                    }
                )
                
                # Convert Amount and Total columns to numeric, handling any errors
                for col in ['Amount', 'Total']:
                    if col in self.historical_df.columns:
                        self.historical_df[col] = pd.to_numeric(
                            self.historical_df[col], 
                            errors='coerce'  # Convert invalid values to NaN
                        ).fillna(0.0)  # Replace NaN with 0.0
                
                logger.info(f"Loaded historical data: {len(self.historical_df)} records")
                
                # Log column types for debugging
                logger.debug("Historical data column types:")
                for col, dtype in self.historical_df.dtypes.items():
                    logger.debug(f"  {col}: {dtype}")
                    
            else:
                logger.info("Historical file not found, creating new DataFrame")
                self._create_empty_historical_df()
                
        except Exception as e:
            logger.error(f"Error loading historical data: {str(e)}")
            self._create_empty_historical_df()
    
    def _create_empty_historical_df(self):
        """Create empty DataFrame with required columns including Order ID."""
        columns = [
            'Release Payment', 'Duplicate Check', 'Full Duplicate Key',
            'Input File', 'Order ID', 'EOBR Number', 'Vendor', 'Mailing Address',
            'Terms', 'Bill Date', 'Due Date', 'Category', 'Description',
            'Amount', 'Memo', 'Total'
        ]
        self.historical_df = pd.DataFrame(columns=columns)
    
    def _init_holidays(self):
        """Initialize US federal holidays for current and next year."""
        current_year = datetime.now().year
        self.us_holidays = holidays.UnitedStates(years=range(current_year, current_year + 2))
        logger.debug(f"Initialized holidays for years {current_year}-{current_year + 1}")
    
    def validate_bill_for_processing(self, bill: Dict[str, Any]) -> Tuple[bool, str]:
        """
        Validate that a bill has all required fields for processing.
        
        Args:
            bill: Bill dictionary
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        required_fields = ['order_id', 'FileMaker_Record_Number', 'PatientName', 'line_items']
        
        for field in required_fields:
            if field not in bill or not bill[field]:
                return False, f"Missing required field: {field}"
        
        # Validate line_items structure
        line_items = bill.get('line_items', [])
        if not isinstance(line_items, list) or len(line_items) == 0:
            return False, "line_items must be a non-empty list"
        
        # Validate each line item has required fields
        for i, item in enumerate(line_items):
            if not isinstance(item, dict):
                return False, f"line_items[{i}] must be a dictionary"
            
            if 'cpt_code' not in item or not item['cpt_code']:
                return False, f"line_items[{i}] missing cpt_code"
            
            if 'allowed_amount' not in item:
                return False, f"line_items[{i}] missing allowed_amount"
        
        return True, ""

    def create_duplicate_key(self, bill: Dict[str, Any]) -> str:
        """
        Create Full Duplicate Key from order_id and CPT codes.
        
        Args:
            bill: Bill dictionary with order_id and line_items
            
        Returns:
            Full Duplicate Key in format: "ORDER_ID|CPT1,CPT2,CPT3"
        """
        try:
            # Get order_id
            order_id = bill.get('order_id', '')
            if not order_id:
                logger.warning(f"No order_id found for bill {bill.get('id')}")
                order_id = 'UNKNOWN'
            
            # Get all CPT codes from line items
            line_items = bill.get('line_items', [])
            cpts = []
            
            for item in line_items:
                cpt = item.get('cpt_code', '').strip()
                if cpt:
                    cpts.append(cpt)
            
            # Sort CPTs for consistent key generation
            cpts.sort()
            cpt_string = ','.join(cpts)
            
            full_key = f"{order_id}|{cpt_string}"
            
            # Log detailed information about key creation
            logger.info(f"Creating duplicate key for bill {bill.get('id')}:")
            logger.info(f"  Order ID: {order_id}")
            logger.info(f"  CPT codes: {cpts}")
            logger.info(f"  Generated key: {full_key}")
            
            return full_key
            
        except Exception as e:
            logger.error(f"Error creating duplicate key for bill {bill.get('id')}: {str(e)}")
            return f"ERROR_{bill.get('id', 'UNKNOWN')}|"

    def enhanced_duplicate_check(self, full_duplicate_key: str, bill: Dict[str, Any]) -> Tuple[bool, str]:
        """
        Enhanced duplicate check using order_id + CPT combinations.
        Only checks for exact duplicates and same order_id with different CPTs.
        
        Args:
            full_duplicate_key: The primary duplicate key to check
            bill: The current bill being processed
            
        Returns:
            Tuple of (is_duplicate, duplicate_type)
            duplicate_type can be: "exact", "same_order_different_cpts", "none"
        """
        if self.historical_df.empty and not self.current_batch_keys:
            return False, "none"
        
        current_order_id = bill.get('order_id', '').strip()
        
        logger.info(f"Enhanced duplicate check for key: {full_duplicate_key}")
        logger.info(f"Current order_id: {current_order_id}")
        
        # Check exact duplicate first (order_id + CPT combination)
        existing_keys = []
        if not self.historical_df.empty and 'Full Duplicate Key' in self.historical_df.columns:
            existing_keys = self.historical_df['Full Duplicate Key'].fillna('').tolist()
        
        # Check historical data for exact match
        historical_exact_match = full_duplicate_key in existing_keys
        if historical_exact_match:
            matching_row = self.historical_df[self.historical_df['Full Duplicate Key'] == full_duplicate_key].iloc[0]
            row_num = self.historical_df[self.historical_df['Full Duplicate Key'] == full_duplicate_key].index[0] + 1
            logger.info(f"Found EXACT duplicate in historical data at row {row_num}:")
            logger.info(f"  EOBR Number: {matching_row['EOBR Number']}")
            logger.info(f"  Order ID: {matching_row.get('Order ID', 'N/A')}")
            logger.info(f"  Description: {matching_row['Description']}")
            
            # Safely handle Amount formatting
            try:
                amount_value = float(matching_row['Amount'])
                logger.info(f"  Amount: ${amount_value:.2f}")
            except (ValueError, TypeError):
                logger.info(f"  Amount: {matching_row['Amount']} (could not format as currency)")
            
            return True, "exact"
        
        # Check current batch for exact match
        current_batch_exact_match = full_duplicate_key in self.current_batch_keys
        if current_batch_exact_match:
            logger.info(f"Found EXACT duplicate in current batch")
            return True, "exact"
        
        # Enhanced check: Look for same order_id with different CPT combinations
        if current_order_id:
            logger.info(f"Checking for same order_id with different CPTs...")
            
            # Check historical data for same order_id
            if not self.historical_df.empty and 'Order ID' in self.historical_df.columns:
                same_order_matches = self.historical_df[
                    self.historical_df['Order ID'] == current_order_id
                ]
                
                if not same_order_matches.empty:
                    logger.warning(f"Found {len(same_order_matches)} records with same order_id but different CPTs:")
                    for idx, match_row in same_order_matches.iterrows():
                        logger.warning(f"  Row {idx + 1}: {match_row.get('Full Duplicate Key', 'N/A')}")
                        logger.warning(f"    EOBR: {match_row.get('EOBR Number', 'N/A')}")
                    
                    # This is a yellow flag - same order but different services
                    return True, "same_order_different_cpts"
        
        # No duplicates found
        logger.debug(f"No duplicates found for: {full_duplicate_key}")
        self.current_batch_keys.add(full_duplicate_key)
        return False, "none"
    
    def get_next_eobr_number(self, fm_record_number: str) -> str:
        """
        Get the next EOBR number for a FileMaker_Record_Number across ALL history.
        
        Args:
            fm_record_number: The FileMaker Record Number
            
        Returns:
            Next EOBR number in format: "FM_RECORD-X"
        """
        try:
            if self.historical_df.empty:
                return f"{fm_record_number}-1"
            
            # Find all existing EOBR numbers for this FM record
            eobr_pattern = f"{re.escape(fm_record_number)}-"
            existing_eobrs = self.historical_df[
                self.historical_df['EOBR Number'].str.contains(eobr_pattern, na=False)
            ]['EOBR Number'].tolist()
            
            if not existing_eobrs:
                next_number = f"{fm_record_number}-1"
                logger.debug(f"First EOBR for {fm_record_number}: {next_number}")
                return next_number
            
            # Extract sequence numbers and find max
            sequence_numbers = []
            for eobr in existing_eobrs:
                try:
                    # Extract number after last dash
                    seq_str = eobr.split('-')[-1]
                    seq = int(seq_str)
                    sequence_numbers.append(seq)
                except (ValueError, IndexError):
                    logger.warning(f"Could not parse sequence from EOBR number: {eobr}")
                    continue
            
            next_seq = max(sequence_numbers) + 1 if sequence_numbers else 1
            next_number = f"{fm_record_number}-{next_seq}"
            
            logger.debug(f"Next EOBR for {fm_record_number}: {next_number} (existing: {len(existing_eobrs)})")
            return next_number
            
        except Exception as e:
            logger.error(f"Error getting EOBR number for {fm_record_number}: {str(e)}")
            return f"{fm_record_number}-1"
    
    def calculate_due_date(self, bill_date_str: str) -> str:
        """
        Calculate due date: bill_date + 45 business days (excluding federal holidays).
        
        Args:
            bill_date_str: Bill date in YYYY-MM-DD format
            
        Returns:
            Due date in YYYY-MM-DD format
        """
        try:
            # Parse bill date
            bill_date = datetime.strptime(bill_date_str, '%Y-%m-%d').date()
            
            current_date = bill_date
            business_days_added = 0
            
            while business_days_added < 45:
                current_date += timedelta(days=1)
                
                # Skip weekends (Saturday=5, Sunday=6)
                if current_date.weekday() >= 5:
                    continue
                
                # Skip federal holidays
                if current_date in self.us_holidays:
                    logger.debug(f"Skipping holiday: {current_date} ({self.us_holidays.get(current_date)})")
                    continue
                
                business_days_added += 1
            
            due_date_str = current_date.strftime('%Y-%m-%d')
            logger.debug(f"Due date calculated: {bill_date_str} + 45 business days = {due_date_str}")
            
            return due_date_str
            
        except Exception as e:
            logger.error(f"Error calculating due date for {bill_date_str}: {str(e)}")
            # Fallback: add 65 calendar days
            try:
                fallback_date = datetime.strptime(bill_date_str, '%Y-%m-%d').date() + timedelta(days=65)
                return fallback_date.strftime('%Y-%m-%d')
            except:
                return bill_date_str
    
    def get_earliest_service_date(self, line_items: List[Dict[str, Any]]) -> str:
        """
        Get the earliest date of service from line items.
        
        Args:
            line_items: List of line item dictionaries
            
        Returns:
            Earliest date in YYYY-MM-DD format
        """
        try:
            dates = []
            
            for item in line_items:
                date_str = item.get('date_of_service', '')
                if not date_str:
                    continue
                
                # Handle date ranges (take first date)
                if ' - ' in date_str:
                    date_str = date_str.split(' - ')[0].strip()
                
                # Try to parse date
                try:
                    # If already in YYYY-MM-DD format
                    if re.match(r'\d{4}-\d{2}-\d{2}', date_str):
                        dates.append(datetime.strptime(date_str, '%Y-%m-%d').date())
                    # If in MM/DD/YYYY format
                    elif re.match(r'\d{1,2}/\d{1,2}/\d{4}', date_str):
                        dates.append(datetime.strptime(date_str, '%m/%d/%Y').date())
                    # If in MM/DD/YY format
                    elif re.match(r'\d{1,2}/\d{1,2}/\d{2}', date_str):
                        dates.append(datetime.strptime(date_str, '%m/%d/%y').date())
                except ValueError as e:
                    logger.warning(f"Could not parse date: {date_str} - {str(e)}")
                    continue
            
            if not dates:
                logger.warning("No valid dates found in line items")
                return datetime.now().strftime('%Y-%m-%d')
            
            earliest_date = min(dates)
            return earliest_date.strftime('%Y-%m-%d')
            
        except Exception as e:
            logger.error(f"Error getting earliest service date: {str(e)}")
            return datetime.now().strftime('%Y-%m-%d')
    
    def format_mailing_address(self, bill: Dict[str, Any]) -> str:
        """
        Format provider mailing address from provider data.
        
        Args:
            bill: Bill dictionary with provider information
            
        Returns:
            Formatted mailing address string
        """
        try:
            address_parts = []
            
            # Billing Address 1
            addr1 = bill.get('provider_billing_address1', '').strip()
            if addr1:
                address_parts.append(addr1)
            
            # Billing Address 2
            addr2 = bill.get('provider_billing_address2', '').strip()
            if addr2:
                address_parts.append(addr2)
            
            # City, State ZIP
            city = bill.get('provider_billing_city', '').strip()
            state = bill.get('provider_billing_state', '').strip()
            zip_code = bill.get('provider_billing_postal_code', '').strip()
            
            city_state_zip_parts = []
            if city:
                city_state_zip_parts.append(city)
            if state:
                city_state_zip_parts.append(state)
            if zip_code:
                city_state_zip_parts.append(zip_code)
            
            if city_state_zip_parts:
                city_state_zip = ', '.join(city_state_zip_parts[:2])  # City, State
                if len(city_state_zip_parts) > 2:  # Add ZIP
                    city_state_zip += f' {city_state_zip_parts[2]}'
                address_parts.append(city_state_zip)
            
            formatted_address = ', '.join(address_parts)
            logger.debug(f"Formatted mailing address: {formatted_address}")
            
            return formatted_address
            
        except Exception as e:
            logger.error(f"Error formatting mailing address: {str(e)}")
            return "Address formatting error"
    
    def calculate_total_amount(self, line_items: List[Dict[str, Any]]) -> float:
        """
        Calculate total allowed amount from line items.
        
        Args:
            line_items: List of line item dictionaries
            
        Returns:
            Total allowed amount
        """
        try:
            total = 0.0
            
            for item in line_items:
                allowed_amount = item.get('allowed_amount')
                if allowed_amount is not None:
                    try:
                        total += float(allowed_amount)
                    except (ValueError, TypeError):
                        logger.warning(f"Could not parse allowed_amount: {allowed_amount}")
                        continue
            
            logger.debug(f"Calculated total amount: ${total:.2f}")
            return total
            
        except Exception as e:
            logger.error(f"Error calculating total amount: {str(e)}")
            return 0.0
    
    def format_description(self, bill: Dict[str, Any], bill_date: str) -> str:
        """
        Format description field: dates, CPTs, patient name, order_id.
        
        Args:
            bill: Bill dictionary
            bill_date: Formatted bill date
            
        Returns:
            Formatted description string
        """
        try:
            description_parts = []
            
            # Add bill date
            description_parts.append(bill_date)
            
            # Add CPT codes
            line_items = bill.get('line_items', [])
            cpts = []
            for item in line_items:
                cpt = item.get('cpt_code', '').strip()
                if cpt:
                    cpts.append(cpt)
            
            if cpts:
                cpts.sort()  # Sort for consistency
                description_parts.append(', '.join(cpts))
            
            # Add patient name
            patient_name = bill.get('PatientName', '').strip()
            if patient_name:
                description_parts.append(patient_name)
            
            # Add Order ID
            order_id = bill.get('order_id', '').strip()
            if order_id:
                description_parts.append(order_id)
            
            description = ', '.join(description_parts)
            logger.debug(f"Formatted description: {description}")
            
            return description
            
        except Exception as e:
            logger.error(f"Error formatting description: {str(e)}")
            return "Description formatting error"
    
    def format_memo(self, bill: Dict[str, Any], bill_date: str) -> str:
        """
        Format memo field: dates, patient name.
        
        Args:
            bill: Bill dictionary
            bill_date: Formatted bill date
            
        Returns:
            Formatted memo string
        """
        try:
            memo_parts = []
            
            # Add bill date
            memo_parts.append(bill_date)
            
            # Add patient name
            patient_name = bill.get('PatientName', '').strip()
            if patient_name:
                memo_parts.append(patient_name)
            
            memo = ', '.join(memo_parts)
            logger.debug(f"Formatted memo: {memo}")
            
            return memo
            
        except Exception as e:
            logger.error(f"Error formatting memo: {str(e)}")
            return "Memo formatting error"
    
    def create_excel_row(self, bill: Dict[str, Any]) -> Dict[str, Any]:
        """
        Create a single Excel row for a bill with order_id-based duplicate detection.
        
        Args:
            bill: Bill dictionary with all required data (must include order_id AND FileMaker_Record_Number)
            
        Returns:
            Dictionary representing one Excel row
        """
        try:
            logger.info(f"Processing bill {bill.get('id')} for Excel generation")
            
            # Validate that we have both order_id and FileMaker_Record_Number
            order_id = bill.get('order_id', '').strip()
            fm_record = bill.get('FileMaker_Record_Number', '').strip()
            
            if not order_id:
                logger.error(f"Bill {bill.get('id')} missing order_id - cannot process")
                raise ValueError(f"Bill {bill.get('id')} missing required order_id")
            
            if not fm_record:
                logger.error(f"Bill {bill.get('id')} missing FileMaker_Record_Number - cannot create EOBR")
                raise ValueError(f"Bill {bill.get('id')} missing required FileMaker_Record_Number")
            
            # Create duplicate key (using order_id)
            full_duplicate_key = self.create_duplicate_key(bill)
            
            # Enhanced duplicate check
            is_duplicate, duplicate_type = self.enhanced_duplicate_check(full_duplicate_key, bill)
            
            # Get EOBR number (using FileMaker_Record_Number)
            eobr_number = self.get_next_eobr_number(fm_record)
            
            # Get earliest service date
            line_items = bill.get('line_items', [])
            bill_date = self.get_earliest_service_date(line_items)
            
            # Calculate due date
            due_date = self.calculate_due_date(bill_date)
            
            # Calculate total amount
            total_amount = self.calculate_total_amount(line_items)
            
            # Format fields
            vendor = bill.get('provider_billing_name', '').strip()
            mailing_address = self.format_mailing_address(bill)
            description = self.format_description(bill, bill_date)
            memo = self.format_memo(bill, bill_date)
            
            # Determine release payment and duplicate check based on duplicate type
            if duplicate_type == "exact":
                release_payment = "N"
                duplicate_check = "Y"
                logger.info(f"Bill marked as EXACT duplicate - will not be released")
            elif duplicate_type == "same_order_different_cpts":
                release_payment = "REVIEW"  # Changed from "Y" to "REVIEW" for manual review
                duplicate_check = "YELLOW"  # Special flag for manual review
                logger.warning(f"Bill marked for manual review - same order_id with different CPTs")
            else:
                release_payment = "Y"
                duplicate_check = "N"
            
            # Create Excel row (INCLUDING ORDER ID)
            excel_row = {
                'Release Payment': release_payment,
                'Duplicate Check': duplicate_check,
                'Full Duplicate Key': full_duplicate_key,
                'Input File': bill.get('id', ''),
                'Order ID': order_id,
                'EOBR Number': eobr_number,  # Still based on FileMaker_Record_Number
                'Vendor': vendor,
                'Mailing Address': mailing_address,
                'Terms': 'Net 45',
                'Bill Date': bill_date,
                'Due Date': due_date,
                'Category': 'Subcontracted Services:Provider Services',
                'Description': description,
                'Amount': round(total_amount, 2),
                'Memo': memo,
                'Total': round(total_amount, 2)
            }
            
            logger.info(f"Created Excel row for bill {bill.get('id')}: "
                       f"Order ID={order_id}, EOBR={eobr_number}, Amount=${total_amount:.2f}, Duplicate={duplicate_check}")
            
            return excel_row
            
        except Exception as e:
            logger.error(f"Error creating Excel row for bill {bill.get('id')}: {str(e)}")
            raise

    def generate_batch_excel(self, 
                        bills: List[Dict[str, Any]], 
                        batch_output_dir: Path,
                        batch_filename: str = "batch_payment_data.xlsx") -> Tuple[Path, int, int, int]:
        """
        Generate batch Excel file and update historical data.
        
        Args:
            bills: List of bill dictionaries
            batch_output_dir: Directory for batch output
            batch_filename: Name of the batch Excel file
            
        Returns:
            Tuple of (batch_excel_path, new_records_count, duplicate_count, yellow_count)
        """
        try:
            logger.info(f"Generating Excel batch for {len(bills)} bills")
            
            # Reset current batch keys for new batch
            self.current_batch_keys.clear()
            
            # Create Excel rows
            excel_rows = []
            duplicate_count = 0
            yellow_count = 0
            
            for bill in bills:
                row = self.create_excel_row(bill)
                excel_rows.append(row)
                
                if row['Duplicate Check'] == 'Y':
                    duplicate_count += 1
                elif row['Duplicate Check'] == 'YELLOW':
                    yellow_count += 1
            
            new_records_count = len(excel_rows) - duplicate_count  # Yellow records are still processed
            
            # Create DataFrame
            batch_df = pd.DataFrame(excel_rows)
            
            # Save batch Excel file with conditional formatting
            excel_dir = batch_output_dir / "excel"
            excel_dir.mkdir(parents=True, exist_ok=True)
            batch_excel_path = excel_dir / batch_filename
            
            # Save with formatting
            self._save_excel_with_formatting(batch_df, batch_excel_path)
            
            logger.info(f"Saved batch Excel file: {batch_excel_path}")
            
            # Update historical data with new records only (excluding exact duplicates)
            new_records = batch_df[batch_df['Duplicate Check'] != 'Y'].copy()
            
            if not new_records.empty:
                # Append to historical DataFrame
                self.historical_df = pd.concat([self.historical_df, new_records], ignore_index=True)
                
                # Save updated historical file
                self.historical_excel_path.parent.mkdir(parents=True, exist_ok=True)
                self.historical_df.to_excel(self.historical_excel_path, index=False)
                logger.info(f"Updated historical file with {len(new_records)} new records")
            else:
                logger.info("No new records to add to historical file")
            
            logger.info(f"Excel generation complete: {len(excel_rows)} total, "
                    f"{new_records_count} new, {duplicate_count} exact duplicates, "
                    f"{yellow_count} same order different CPTs warnings")
            
            return batch_excel_path, new_records_count, duplicate_count, yellow_count
            
        except Exception as e:
            logger.error(f"Error generating batch Excel: {str(e)}")
            raise

    def _save_excel_with_formatting(self, df: pd.DataFrame, file_path: Path):
        """
        Save Excel file with conditional formatting for yellow warnings.
        
        Args:
            df: DataFrame to save
            file_path: Path to save the Excel file
        """
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Batch')
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Batch']
                
                # Add conditional formatting for yellow warnings
                try:
                    from openpyxl.styles import PatternFill
                    
                    # Yellow fill for YELLOW duplicate check
                    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    
                    # Apply to rows where Duplicate Check = "YELLOW"
                    duplicate_check_col = None
                    for col_idx, col_name in enumerate(df.columns, 1):
                        if col_name == 'Duplicate Check':
                            duplicate_check_col = col_idx
                            break
                    
                    if duplicate_check_col:
                        for row_num, (_, row_data) in enumerate(df.iterrows(), start=2):  # Start at 2 to skip header
                            if row_data['Duplicate Check'] == 'YELLOW':
                                for col_num in range(1, len(df.columns) + 1):
                                    worksheet.cell(row=row_num, column=col_num).fill = yellow_fill
                    
                    logger.info(f"Applied conditional formatting to {len(df[df['Duplicate Check'] == 'YELLOW'])} yellow rows")
                    
                except ImportError:
                    logger.warning("openpyxl not available for conditional formatting")
                except Exception as e:
                    logger.warning(f"Could not apply conditional formatting: {str(e)}")
        
        except Exception as e:
            # Fallback to basic save if formatting fails
            logger.warning(f"Conditional formatting failed, saving basic Excel: {str(e)}")
            df.to_excel(file_path, index=False)

    def get_batch_summary(self, batch_excel_path: Path) -> Dict[str, Any]:
        """
        Get summary information about a generated batch with enhanced duplicate info.
        
        Args:
            batch_excel_path: Path to the batch Excel file
            
        Returns:
            Dictionary with batch summary information
        """
        try:
            batch_df = pd.read_excel(batch_excel_path)
            
            summary = {
                'total_records': len(batch_df),
                'new_records': len(batch_df[batch_df['Duplicate Check'] == 'N']),
                'duplicate_records': len(batch_df[batch_df['Duplicate Check'] == 'Y']),
                'yellow_warnings': len(batch_df[batch_df['Duplicate Check'] == 'YELLOW']),
                'total_amount': batch_df['Amount'].sum(),
                'release_amount': batch_df[batch_df['Release Payment'] == 'Y']['Amount'].sum(),
                'batch_file': str(batch_excel_path),
                'generation_time': datetime.now().isoformat()
            }
            
            return summary
            
        except Exception as e:
            logger.error(f"Error getting batch summary: {str(e)}")
            return {}

def generate_excel_batch(bills: List[Dict[str, Any]], 
                        batch_output_dir: Path,
                        historical_excel_path: Path = None) -> Tuple[Path, Dict[str, Any]]:
    """
    Convenience function to generate Excel batch and return summary with enhanced duplicate detection.
    
    Args:
        bills: List of prepared bill dictionaries
        batch_output_dir: Directory for batch output
        historical_excel_path: Path to historical Excel file
        
    Returns:
        Tuple of (batch_excel_path, summary_dict)
    """
    # Set up logging for this run
    logging.basicConfig(level=logging.INFO)
    logger.setLevel(logging.INFO)
    
    generator = ExcelBatchGenerator(historical_excel_path)
    batch_excel_path, new_count, dup_count, yellow_count = generator.generate_batch_excel(bills, batch_output_dir)
    summary = generator.get_batch_summary(batch_excel_path)
    
    return batch_excel_path, summary

if __name__ == "__main__":
    # Test the Excel generator
    import logging
    
    logging.basicConfig(level=logging.INFO)
    
    # Sample bill data for testing
    test_bills = [
        {
            'id': 'BILL_001',
            'order_id': '9A7546EF-D14D-46B2-8EE8-AB16255B4F12',
            'FileMaker_Record_Number': 'FM_12345',
            'PatientName': 'John Doe',
            'provider_billing_name': 'Test Medical Center',
            'provider_billing_address1': '123 Medical Drive',
            'provider_billing_city': 'Orlando',
            'provider_billing_state': 'FL',
            'provider_billing_postal_code': '32801',
            'line_items': [
                {
                    'cpt_code': '99213',
                    'allowed_amount': 120.00,
                    'date_of_service': '2024-01-15'
                },
                {
                    'cpt_code': '73610',
                    'allowed_amount': 180.00,
                    'date_of_service': '2024-01-15'
                }
            ]
        }
    ]
    
    # Test generation
    try:
        output_dir = Path("test_batch")
        batch_path, summary = generate_excel_batch(test_bills, output_dir)
        print(f"✅ Test Excel generated: {batch_path}")
        print(f"📊 Summary: {summary}")
    except Exception as e:
        print(f"❌ Error: {str(e)}")